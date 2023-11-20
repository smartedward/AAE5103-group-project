import os
from datetime import datetime
import cv2
import numpy as np
from openpyxl import Workbook
import utils.utils as utils
#from utils import utils
from net.inception import InceptionResNetV1
from net.mtcnn import mtcnn


class face_rec():
    def __init__(self):
        #-------------------------#
        #   创建mtcnn的模型
        #   用于检测人脸
        #-------------------------#
        self.mtcnn_model = mtcnn()
        self.threshold = [0.5,0.6,0.8]
               
        #-----------------------------------#
        #   载入facenet
        #   将检测到的人脸转化为128维的向量
        #-----------------------------------#
        self.facenet_model = InceptionResNetV1()
        model_path = './model_data/facenet_keras.h5'
        self.facenet_model.load_weights(model_path)

        #-----------------------------------------------#
        #   对数据库中的人脸进行编码
        #   known_face_encodings中存储的是编码后的人脸
        #   known_face_names为人脸的名字
        #-----------------------------------------------#
        face_list = os.listdir("face_dataset")
        self.known_face_encodings=[]
        self.known_face_names=[]
        self.known_face_airlines = []
        self.known_face_genders = []
        self.known_face_satisfactions = []
        for face in face_list:
            name = face.split(".")[0]
            airline = face.split(".")[1]
            gender = face.split(".")[2]
            satisfaction = face.split(".")[3]
            img = cv2.imread("./face_dataset/"+face)
            img = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
            #---------------------#
            #   检测人脸
            #---------------------#
            rectangles = self.mtcnn_model.detectFace(img, self.threshold)
            #---------------------#
            #   转化成正方形
            #---------------------#
            rectangles = utils.rect2square(np.array(rectangles))
            #-----------------------------------------------#
            #   facenet要传入一个160x160的图片
            #   利用landmark对人脸进行矫正
            #-----------------------------------------------#
            rectangle = rectangles[0]
            landmark = np.reshape(rectangle[5:15], (5,2)) - np.array([int(rectangle[0]), int(rectangle[1])])
            crop_img = img[int(rectangle[1]):int(rectangle[3]), int(rectangle[0]):int(rectangle[2])]
            crop_img, _ = utils.Alignment_1(crop_img,landmark)
            crop_img = np.expand_dims(cv2.resize(crop_img, (160, 160)), 0)
            #--------------------------------------------------------------------#
            #   将检测到的人脸传入到facenet的模型中，实现128维特征向量的提取
            #--------------------------------------------------------------------#
            face_encoding = utils.calc_128_vec(self.facenet_model, crop_img)

            self.known_face_encodings.append(face_encoding)
            self.known_face_names.append(name)
            self.known_face_airlines.append(airline)
            self.known_face_genders.append(gender)
            self.known_face_satisfactions.append(satisfaction)

    def recognize(self,draw):
        #-----------------------------------------------#
        #   人脸识别
        #   先定位，再进行数据库匹配
        #-----------------------------------------------#
        # draw = cv2.resize(odraw,(800,800))
        height,width,_ = np.shape(draw)
        draw_rgb = cv2.cvtColor(draw,cv2.COLOR_BGR2RGB)

        #--------------------------------#
        #   检测人脸
        #--------------------------------#
        rectangles = self.mtcnn_model.detectFace(draw_rgb, self.threshold)

        if len(rectangles)==0:
            return

        # 转化成正方形
        rectangles = utils.rect2square(np.array(rectangles,dtype=np.int32))
        rectangles[:, [0,2]] = np.clip(rectangles[:, [0,2]], 0, width)
        rectangles[:, [1,3]] = np.clip(rectangles[:, [1,3]], 0, height)

        #-----------------------------------------------#
        #   对检测到的人脸进行编码
        #-----------------------------------------------#
        face_encodings = []
        for rectangle in rectangles:
            #---------------#
            #   截取图像
            #---------------#
            landmark = np.reshape(rectangle[5:15], (5,2)) - np.array([int(rectangle[0]), int(rectangle[1])])
            crop_img = draw_rgb[int(rectangle[1]):int(rectangle[3]), int(rectangle[0]):int(rectangle[2])]
            #-----------------------------------------------#
            #   利用人脸关键点进行人脸对齐
            #-----------------------------------------------#
            crop_img,_ = utils.Alignment_1(crop_img,landmark)
            crop_img = np.expand_dims(cv2.resize(crop_img, (160, 160)), 0)

            face_encoding = utils.calc_128_vec(self.facenet_model, crop_img)
            face_encodings.append(face_encoding)

        face_names = []
        face_airlines =[]
        face_genders= []
        face_satisfactions=[]
        for face_encoding in face_encodings:
            #-------------------------------------------------------#
            #   取出一张脸并与数据库中所有的人脸进行对比，计算得分
            #-------------------------------------------------------#
            matches = utils.compare_faces(self.known_face_encodings, face_encoding, tolerance = 0.9)
            name = "Unknown"
            airline="Unknown"
            gender="Unknown"
            satisfaction="Unknown"
            #-------------------------------------------------------#
            #   找出距离最近的人脸
            #-------------------------------------------------------#
            face_distances = utils.face_distance(self.known_face_encodings, face_encoding)
            #-------------------------------------------------------#
            #   取出这个最近人脸的评分
            #-------------------------------------------------------#
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = self.known_face_names[best_match_index]
                airline = self.known_face_airlines[best_match_index]
                gender = self.known_face_genders[best_match_index]
                satisfaction = self.known_face_satisfactions[best_match_index]
            face_names.append(name)
            face_airlines.append(airline)
            face_genders.append(gender)
            face_satisfactions.append(satisfaction)


        rectangles = rectangles[:,0:4]
        #-----------------------------------------------#
        #   画框~!~
        #-----------------------------------------------#
        for (left, top, right, bottom), name, airline, gender,satisfaction in zip(rectangles, face_names,face_airlines,face_genders,face_satisfactions):
            cv2.rectangle(draw, (left, top), (right, bottom), (255, 255, 255), 2)
            font = cv2.FONT_HERSHEY_SIMPLEX
            cv2.putText(draw, name, (left , top+15), font, 0.75, (255, 255, 255), 2)
            cv2.putText(draw, airline, (left, top+40), font, 0.75, (255, 255, 255), 2)
            cv2.putText(draw, gender, (left, top + 65), font, 0.75, (255, 255, 255), 2)
            # draw=cv2.resize(draw,(800,400))
            self.detect_if_success = matches[best_match_index]
            self.name=name
            self.airline=airline
            self.gender=gender
            self.satisfaction=satisfaction
            print(matches)
        return draw

    #def detect_if_success(self):
        #normal_bool=self.detect_if_success().item()
        #return self.detect_if_success
    def create_passengerfile(self):
        # 创建一个工作簿
        wb = Workbook()
        # 选择默认的工作表（第一个工作表）
        ws = wb.active
        # 获取当前的日期和时间
        current_datetime = datetime.now()
        # 在工作表中写入信息
        ws['A1'] = '姓名'
        ws['A2'] = '性别'
        ws['A3'] = '航班'
        ws['A5'] = '录入信息时间'
        ws['A4'] = '国籍'
        ws['B1'] = self.name
        ws['B2'] = self.gender
        ws['B3'] = self.airline
        ws['B5'] = current_datetime
        ws['B4'] = self.satisfaction
        # 指定保存的文件夹路径
        folder_path = 'D:/pycharm/opencv/keras-face-recognition-master/passengers_files'
        # 确保文件夹存在，如果不存在就创建它
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        # 拼接文件的完整路径
        file_path = os.path.join(folder_path, self.name+'.xlsx')
        # 保存工作簿到指定的文件路径
        wb.save(file_path)
        print(f'Excel文件已保存到: {file_path}')

    def count_files_in_folder(self,folder_path):
        # 确保路径存在且是一个文件夹
        if os.path.exists(folder_path) and os.path.isdir(folder_path):
            # 获取文件夹中的所有文件列表
            files = os.listdir(folder_path)
            # 使用列表推导式过滤出文件（排除文件夹）
            files = [f for f in files if os.path.isfile(os.path.join(folder_path, f))]
            # 统计文件数量
            num_files = len(files)
        return num_files

if __name__ == "__main__":
    ai_assignment = face_rec()
    j = ai_assignment.count_files_in_folder(f"D:/pycharm/opencv/keras-face-recognition-master/captured vedios")
    for i in range(j):
        videos_path = f"D:/pycharm/opencv/keras-face-recognition-master/captured vedios/{i + 1}.mp4"
        video_capture = cv2.VideoCapture(videos_path)
        # 创建一个窗口并设置其大小
        cv2.namedWindow('detect', cv2.WINDOW_NORMAL)
        cv2.resizeWindow('detect', 900, 600)
        while True:
            ret, draw = video_capture.read()
            ai_assignment.recognize(draw)
            cv2.imshow('detect', draw)
            cv2.waitKey(2000)
            ai_assignment.create_passengerfile()
            if cv2.waitKey(1) & 0xFF == ord(' ') or ai_assignment.detect_if_success:
                break
        video_capture.release()
        cv2.destroyAllWindows()

